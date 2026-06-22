from __future__ import annotations

import io
import json
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.worksheet import Worksheet


def sanitize_excel_filename(product_name: str | None) -> str:
    """Create stable workbook filename from product name.

    Input: optional product name from Agent 1 output.
    Output: `<ProductName>_PRD_Analysis.xlsx` with safe characters only.
    """

    base = re.sub(r"[^A-Za-z0-9]+", "", str(product_name or "").strip()) or "PRD"
    return f"{base}_PRD_Analysis.xlsx"


def build_prd_excel_bytes(result: dict[str, Any]) -> tuple[str, bytes]:
    """Generate an Excel workbook from final PRD pipeline result.

    Input: pipeline result dictionary returned by `run_prd_pipeline`.
    Output: tuple of (`filename`, workbook bytes).
    """

    agent_1 = result.get("agent_1_prd_structure") or result.get("agent_1") or {}
    agent_2 = result.get("agent_2_review") or {}
    agent_3 = result.get("agent_3_test_design") or {}
    agent_4 = result.get("agent_4_automation_design") or {}
    audit_log = result.get("audit_log") or []

    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_rows(
        workbook.create_sheet("Summary"),
        ["Field", "Value"],
        [
            ["Status", result.get("status", "")],
            ["Product Name", agent_1.get("product_name", "")],
            ["Business Goal", agent_1.get("business_goal", "")],
            ["Review Loops", result.get("review_loops", "")],
            ["Started At", result.get("started_at", "")],
            ["Completed At", result.get("completed_at", "")],
            ["Provider", result.get("provider", "")],
            ["Cloud Model", result.get("cloud_model", "")],
        ],
    )

    _write_rows(
        workbook.create_sheet("Personas"),
        ["Name", "Goal"],
        [[item.get("name", ""), item.get("goal", "")] for item in (agent_1.get("personas") or [])],
    )

    epics = agent_1.get("epics") or []
    _write_rows(
        workbook.create_sheet("Epics"),
        ["Epic ID", "Title", "Description", "Source Requirement", "Target Persona"],
        [
            [
                epic.get("id", ""),
                epic.get("title", ""),
                epic.get("description", ""),
                epic.get("source_requirement", ""),
                epic.get("target_persona", ""),
            ]
            for epic in epics
        ],
    )

    user_story_rows: list[list[Any]] = []
    acceptance_rows: list[list[Any]] = []
    for epic in epics:
        epic_id = epic.get("id", "")
        epic_title = epic.get("title", "")
        for story in epic.get("user_stories") or []:
            user_story_rows.append(
                [
                    epic_id,
                    epic_title,
                    story.get("id", ""),
                    story.get("title", ""),
                    story.get("story", ""),
                    story.get("priority", ""),
                ]
            )
            for criterion in story.get("acceptance_criteria") or []:
                acceptance_rows.append([epic_id, story.get("id", ""), story.get("title", ""), criterion])

    _write_rows(
        workbook.create_sheet("User Stories"),
        ["Epic ID", "Epic Title", "Story ID", "Story Title", "Story", "Priority"],
        user_story_rows,
    )
    _write_rows(
        workbook.create_sheet("Acceptance Criteria"),
        ["Epic ID", "Story ID", "Story Title", "Acceptance Criterion"],
        acceptance_rows,
    )

    qa_rows: list[list[Any]] = []
    for finding_type in [
        "blockers",
        "missing_requirements",
        "ambiguous_requirements",
        "missing_acceptance_criteria",
        "testability_concerns",
    ]:
        for finding in agent_2.get(finding_type) or []:
            qa_rows.append([agent_2.get("decision", ""), finding_type, finding])
    _write_rows(
        workbook.create_sheet("QA Review"),
        ["Decision", "Finding Type", "Finding"],
        qa_rows,
    )

    _write_rows(
        workbook.create_sheet("Story Tests"),
        ["Story ID", "Title", "Test Type", "Priority", "Preconditions", "Steps", "Expected Result"],
        [
            [
                test.get("story_id", ""),
                test.get("title", ""),
                test.get("test_type", ""),
                test.get("priority", ""),
                test.get("preconditions", ""),
                test.get("steps", ""),
                test.get("expected_result", ""),
            ]
            for test in (agent_3.get("story_level_tests") or [])
        ],
    )

    _write_rows(
        workbook.create_sheet("Epic Tests"),
        ["Epic ID", "Title", "Priority", "Preconditions", "Steps", "Expected Result"],
        [
            [
                test.get("epic_id", ""),
                test.get("title", ""),
                test.get("priority", ""),
                test.get("preconditions", ""),
                test.get("steps", ""),
                test.get("expected_result", ""),
            ]
            for test in (agent_3.get("epic_level_tests") or [])
        ],
    )

    _write_rows(
        workbook.create_sheet("Traceability Matrix"),
        ["Epic ID", "Story ID", "Coverage"],
        [
            [item.get("epic_id", ""), item.get("story_id", ""), item.get("coverage", "")]
            for item in (agent_3.get("traceability_matrix") or [])
        ],
    )

    _write_rows(
        workbook.create_sheet("Automation Suites"),
        ["Name", "Framework", "Test Type", "Steps", "Locator Strategy", "Alumnium Optional Steps"],
        [
            [
                suite.get("name", ""),
                suite.get("framework", ""),
                suite.get("test_type", ""),
                "\n".join(suite.get("steps") or []),
                suite.get("locator_strategy", ""),
                "\n".join(suite.get("alumnium_optional_steps") or []),
            ]
            for suite in (agent_4.get("automation_suites") or [])
        ],
    )

    _write_rows(
        workbook.create_sheet("CI Notes"),
        ["Note"],
        [[note] for note in (agent_4.get("ci_notes") or [])],
    )

    _write_rows(
        workbook.create_sheet("Audit Log"),
        ["Step", "Agent", "Decision", "Loop", "Details"],
        [
            [
                index,
                entry.get("agent", ""),
                entry.get("decision", ""),
                entry.get("loop", ""),
                json.dumps(entry, ensure_ascii=False),
            ]
            for index, entry in enumerate(audit_log, start=1)
        ],
    )

    output = io.BytesIO()
    workbook.save(output)
    return sanitize_excel_filename(agent_1.get("product_name")), output.getvalue()


def _write_rows(sheet: Worksheet, headers: list[str], rows: list[list[Any]]) -> None:
    """Apply lightweight, consistent worksheet formatting and write data rows."""

    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    header_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for cell in sheet[1]:
        cell.font = header_font

    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None:
                cell.alignment = wrap_alignment

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for column in sheet.columns:
        max_len = 0
        for cell in column:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 12), 80)

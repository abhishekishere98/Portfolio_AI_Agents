from __future__ import annotations

import io
import unittest

from openpyxl import load_workbook

from prd_excel_export import build_prd_excel_bytes, sanitize_excel_filename


class PrdExcelExportTests(unittest.TestCase):
    def test_sanitize_excel_filename(self):
        self.assertEqual("SmartWishlist_PRD_Analysis.xlsx", sanitize_excel_filename("Smart Wishlist"))

    def test_build_prd_excel_bytes_contains_expected_sheets(self):
        result = {
            "status": "APPROVED",
            "provider": "cloud",
            "cloud_model": "gemini-2.5-flash-lite",
            "started_at": "2026-06-22T00:00:00Z",
            "completed_at": "2026-06-22T00:01:00Z",
            "review_loops": 1,
            "agent_1_prd_structure": {
                "product_name": "Smart Wishlist",
                "business_goal": "Increase retention",
                "personas": [{"name": "Buyer", "goal": "Save products"}],
                "epics": [
                    {
                        "id": "EPIC-1",
                        "title": "Wishlist Core",
                        "description": "Manage saved products",
                        "source_requirement": "User can add and remove products",
                        "target_persona": "Buyer",
                        "user_stories": [
                            {
                                "id": "US-1.1",
                                "title": "Add item",
                                "story": "As a buyer, I want to add items",
                                "priority": "High",
                                "acceptance_criteria": ["Given product page when add clicked then item is saved"],
                            }
                        ],
                    }
                ],
            },
            "agent_2_review": {
                "decision": "APPROVED",
                "blockers": [],
                "missing_requirements": [],
                "ambiguous_requirements": ["Clarify guest wishlist merge"],
                "missing_acceptance_criteria": [],
                "testability_concerns": [],
            },
            "agent_3_test_design": {
                "story_level_tests": [
                    {
                        "story_id": "US-1.1",
                        "title": "Add wishlist item",
                        "test_type": "Positive",
                        "priority": "High",
                        "preconditions": "Logged in",
                        "steps": "Open product and click wishlist",
                        "expected_result": "Item appears in wishlist",
                    }
                ],
                "epic_level_tests": [
                    {
                        "epic_id": "EPIC-1",
                        "title": "Wishlist E2E",
                        "priority": "High",
                        "preconditions": "Data available",
                        "steps": "Add and remove item",
                        "expected_result": "State updates correctly",
                    }
                ],
                "traceability_matrix": [{"epic_id": "EPIC-1", "story_id": "US-1.1", "coverage": "full"}],
            },
            "agent_4_automation_design": {
                "automation_suites": [
                    {
                        "name": "Wishlist Suite",
                        "framework": "playwright",
                        "test_type": "e2e",
                        "steps": ["Login", "Add item"],
                        "locator_strategy": "data-testid",
                        "alumnium_optional_steps": ["al.do('add item')"],
                    }
                ],
                "ci_notes": ["Run nightly"],
            },
            "audit_log": [{"agent": "prd_analyst", "decision": "VALID_PRD"}],
        }

        filename, workbook_bytes = build_prd_excel_bytes(result)
        self.assertEqual("SmartWishlist_PRD_Analysis.xlsx", filename)
        self.assertGreater(len(workbook_bytes), 0)

        workbook = load_workbook(io.BytesIO(workbook_bytes))
        self.assertEqual(
            [
                "Summary",
                "Personas",
                "Epics",
                "User Stories",
                "Acceptance Criteria",
                "QA Review",
                "Story Tests",
                "Epic Tests",
                "Traceability Matrix",
                "Automation Suites",
                "CI Notes",
                "Audit Log",
            ],
            workbook.sheetnames,
        )


if __name__ == "__main__":
    unittest.main()
